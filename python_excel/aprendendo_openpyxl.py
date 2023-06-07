import openpyxl as xl

# mudando pasta de trabalho
import os 
os.chdir("python_excel")

# criando objeto que é a planilha
planilha  = xl.Workbook()

# cria folha
planilha.create_sheet("folha_1")

# acessar folha
folha_1 = planilha["folha_1"]

# acessar células
folha_1["A1"] = "Data"

# criar tabela
folha_1["B1"] = "Valor"
folha_1["C1"] = "Descrição"

# adicionando uma linha de dados no fim do arquivo
dados = ["10/10/2020", 900, "Aluguel"]

folha_1.append(dados)

# adicionando várias linhas de dados
dados = [["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"]]

for registro in dados:
    folha_1.append(registro)

# criando nova folha
planilha.create_sheet("folha_2")
folha_2 = planilha["folha_2"]

# acessando célula por linha e coluna
# como um objeto
folha_2.cell(row = 5, column = 5).value = "Data"
folha_2.cell(row = 5, column = 6).value = "Valor"
folha_2.cell(row = 5, column = 7).value = "Descrição"

# adicionando dados na folha 2
dados = [["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"],
         ["10/10/2020", 900, "Aluguel"]]

# definindo começo da tabela
# comeco = (5,5)
# pegando útlima linha que foi acrescentada para definir o começo de onde os dados serão colocadas
comeco = (folha_2.max_row + 1, folha_2.min_column)

for linha, dado in enumerate(dados):
    for coluna, valor in enumerate(dado):
        folha_2.cell(row = linha + comeco[0], column = coluna + comeco[1]).value = valor


# salvando planilha
planilha.save("monitoriazinha_cria.xlsx")