import openpyxl as xl

# mudando pasta de trabalho
from os import chdir
chdir("python_excel")

planilha = xl.load_workbook("aula8.xlsx")

folha_1 = planilha["Folha 1"]

# hora de preencher a planilha já criada

#dados
dados = []

# primeiro, precisamos saber onde começa e termina nossa tabela
range_tabela = {"p_linha" : folha_1.min_row, "p_coluna" : folha_1.min_column, "u_minima" :folha_1.max_row, "u_coluna" : folha_1.max_column}

# for linha in range(range_tabela["p_linha"], range_tabela["u_linha"] + 1):
#     for coluna in range(range_tabela["p_coluna"], range_tabela["u_coluna"] + 1):